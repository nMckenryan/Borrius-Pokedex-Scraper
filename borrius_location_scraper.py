import json
import openpyxl
from termcolor import colored

from helpers import correct_pokemon_name

wb = openpyxl.load_workbook("./scraperData/borrius_location_data.xlsx", data_only=True)


async def get_grasscave_locations(locationDataList):
    """
    Grabs all the Pokemon and their locations from the Grass & Cave Encounters sheet of the spreadsheet.
    Applies the headers as locationData.location and the values as "pokemon" to each locationData object in the locationDataList.
    If the Pokemon already exists in locationDataList, it appends the locationData object to the existing Pokemon's locationData array.
    """
    try:
        grassSheet = wb["Grass & Cave Encounters"]
        # map through each column of the spreadsheet, apply the headers as locationData.location
        for col in range(1, grassSheet.max_column + 1):
            location_header = grassSheet.cell(row=1, column=col).value
            isSpecialEncounter = 0
            # map down through each row, applying each value from there as "pokemon"
            for row in range(2, grassSheet.max_row + 1):
                pokemon = grassSheet.cell(row=row, column=col).value
                if pokemon is None:
                    continue
                pokemonName = correct_pokemon_name(pokemon)

                # Special encounters are always shown at the bottom of the row.
                if pokemonName == "Special Encounter":
                    isSpecialEncounter = 1
                    continue
                fontColor = grassSheet.cell(row=row, column=col).font.color
                if fontColor:
                    rgb = [0, 0, 0]
                    rgb = f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
                    match rgb:
                        # orange
                        case "#FFB500":
                            timeOfDay = "Morning"
                        # purple
                        case "#B300FF":
                            timeOfDay = "Night"
                        # red
                        case "#FF0000":
                            timeOfDay = "Swarm"
                        case "#000000":
                            timeOfDay = "All Day"
                        case _:
                            timeOfDay = "Other"
                else:
                    timeOfDay = "Other"

                # check if "pokemon" already exists in locationDataList, then appends locationData object to array
                existingPokemon = next(
                    (p for p in locationDataList if p["pokemon"] == pokemonName), None
                )
                if existingPokemon is not None:
                    existingPokemon["locationData"].append(
                        {
                            "location": location_header,
                            "encounterMethod": "Grass/Cave",
                            "timeOfDay": timeOfDay,
                            "isSpecialEncounter": isSpecialEncounter,
                        }
                    )
                else:
                    locationDataList.append(
                        {
                            "pokemon": pokemonName,
                            "locationData": [
                                {
                                    "location": location_header,
                                    "encounterMethod": "Grass/Cave",
                                    "timeOfDay": timeOfDay,
                                    "isSpecialEncounter": isSpecialEncounter,
                                }
                            ],
                        }
                    )

    except Exception as e:
        print(
            colored(
                f"Failed to get Grass/Cave Locations: {e}",
                "red",
            ),
        )


# Fish Locations also includes rocksmash
async def get_fishing_locations(locationDataList):
    """
    Retrieves all the fishing and rock smash locations from the spreadsheet.

    This function also applies the correct encounter method, time of day, and special encounter status to each pokemon entry.
    """
    try:
        grassSheet = wb["fishingRockSmash"]
        # map through each column of the spreadsheet, apply the headers as locationData.location
        for col in range(1, grassSheet.max_column + 1):
            location_header = grassSheet.cell(row=1, column=col).value
            isSpecialEncounter = 0

            isGoodRod = 0
            isSuperRod = 0
            isUnderwater = 0
            isRockSmash = 0

            # map down through each row, applying each value from there as "pokemon"
            for row in range(2, grassSheet.max_row + 1):
                pokemon = grassSheet.cell(row=row, column=col).value
                if pokemon is None:
                    continue

                # Special encounters are always shown at the bottom of the row (only one in this case).
                if pokemon == "Special Encounter":
                    isSpecialEncounter = 1
                    isGoodRod = 0
                    isSuperRod = 0
                    isUnderwater = 1
                    isRockSmash = 0
                    continue

                if pokemon == "Good Rod":
                    isGoodRod = 1
                    isSuperRod = 0
                    isUnderwater = 0
                    isRockSmash = 0
                    continue
                elif pokemon == "Super Rod":
                    isGoodRod = 0
                    isSuperRod = 1
                    isUnderwater = 0
                    isRockSmash = 0
                elif pokemon == "Underwater":
                    isGoodRod = 0
                    isSuperRod = 0
                    isUnderwater = 1
                    isRockSmash = 0
                    continue
                elif pokemon == "Rock Smash":
                    isGoodRod = 0
                    isSuperRod = 0
                    isUnderwater = 0
                    isRockSmash = 1
                    continue

                method = "Unknown"

                match (isGoodRod, isSuperRod, isUnderwater, isRockSmash):
                    case (1, 0, 0, 0):
                        method = "Good Rod"
                    case (0, 1, 0, 0):
                        method = "Super Rod"
                    case (0, 0, 1, 0):
                        method = "Underwater"
                    case (0, 0, 0, 1):
                        method = "Rock Smash"
                    case _:
                        method = "Unknown"

                pokemonName = correct_pokemon_name(pokemon)
                # check if "pokemon" already exists in locationDataList, then appends locationData object to array
                existingPokemon = next(
                    (p for p in locationDataList if p["pokemon"] == pokemonName), None
                )
                if existingPokemon is not None:
                    existingPokemon["locationData"].append(
                        {
                            "location": location_header,
                            "encounterMethod": method,
                            "timeOfDay": "All Day",
                            "isSpecialEncounter": isSpecialEncounter,
                        }
                    )
                else:
                    locationDataList.append(
                        {
                            "pokemon": pokemonName,
                            "locationData": [
                                {
                                    "location": location_header,
                                    "encounterMethod": method,
                                    "timeOfDay": "All Day",
                                    "isSpecialEncounter": isSpecialEncounter,
                                }
                            ],
                        }
                    )

    except Exception as e:
        print(
            colored(
                f"Failed to retrieve fishing data {e}",
                "red",
            ),
        )


# SURF LOCATIONS
async def get_surf_locations(locationDataList):
    """
    Retrieves all the surfing locations from the spreadsheet.

    This function also applies the correct encounter method, time of day, and special encounter status to each pokemon entry.
    """
    try:
        surfSheet = wb["surfing"]
        # map through each column of the spreadsheet, apply the headers as locationData.location
        for col in range(1, surfSheet.max_column + 1):
            location_header = surfSheet.cell(row=1, column=col).value
            isSpecialEncounter = 0
            # map down through each row, applying each value from there as "pokemon"
            for row in range(2, surfSheet.max_row + 1):
                pokemon = surfSheet.cell(row=row, column=col).value
                if pokemon is None:
                    continue

                # Special encounters are always shown at the bottom of the row.
                if pokemon == "Special Encounter":
                    isSpecialEncounter = 1
                    continue

                # check if "pokemon" already exists in locationDataList, then appends locationData object to array
                existingPokemon = next(
                    (p for p in locationDataList if p["pokemon"] == pokemon), None
                )
                
                if existingPokemon is not None:
                    existingPokemon["locationData"].append(
                        {
                            "location": location_header,
                            "encounterMethod": "Surfing",
                            "timeOfDay": "All Day",
                            "isSpecialEncounter": isSpecialEncounter,
                        }
                    )
                else:
                    locationDataList.append(
                        {
                            "pokemon": pokemon,
                            "locationData": [
                                {
                                    "location": location_header,
                                    "encounterMethod": "Surfing",
                                    "timeOfDay": "All Day",
                                    "isSpecialEncounter": isSpecialEncounter,
                                }
                            ],
                        }
                    )

    except Exception as e:
        print(
            colored(
                f"Failed to get Surf Locations: {e}",
                "red",
            ),
        )


def fill_in_evolution_gaps(locationDataList):
    """
    Fills in any missing evolution data by grabbing the locations from pokelocation.json and applying them to the pokemon's locationData if it exists and is currently empty.

    This function is necessary because some evolutions only have a location listed in the pokelocation.json file, and not in the borrius spreadsheet.
    """

    try:
        with open("scraperData/pokelocation.json", "r", encoding="utf-8") as json_file:
            data = json.load(json_file)
            for pokemon in data:
                pokemonName = correct_pokemon_name(pokemon["name"])

                existingPokemon = next(
                    (p for p in locationDataList if p["pokemon"] == pokemonName), None
                )

                if (
                    existingPokemon is not None
                    and len(existingPokemon["locationData"]) == 0
                ):
                    existingPokemon["locationData"].append(
                        {
                            "location": pokemon["location"],
                            "encounterMethod": "Evolution",
                            "timeOfDay": "All Day",
                            "isSpecialEncounter": 0,
                        }
                    )

    except Exception as e:
        print(
            colored(
                f"fill_in_evolution_gaps failed: {e}",
                "red",
            ),
        )

async def print_location_json():
    """
    Combines data from borrius_pokedex_data.json and pokelocation.json to create a single json file containing all the data from both files.

    The function first loads the borrius_pokedex_data.json file and uses it to populate the locationDataList with pokemon names that may not exist in the pokelocation.json file.

    It then runs the get_grasscave_locations, get_surf_locations, and get_fishing_locations functions to fill in the location data for each pokemon.

    The fill_in_evolution_gaps function is then run to fill in any missing location data for pokemon that evolve from other pokemon.

    Finally, the function writes the locationDataList to a json file named locationData.json in the scraperData directory.

    The function prints a success message if the json file is successfully created, and an error message if the json file generation fails.
    """
    locationDataList = []
    try:
        # await get_full_borrius_pokemon_names(locationDataList)
        await get_grasscave_locations(locationDataList)
        await get_surf_locations(locationDataList)
        await get_fishing_locations(locationDataList)
        fill_in_evolution_gaps(locationDataList)
        
        print_json_file(locationDataList)
        
        print(
            colored(
                "locationData.json successfully created",
                "green",
            ),
        )
    except Exception as e:
        print(
            colored(
                f"locationData Json Generation Failed : {e}",
                "red",
            ),
        )


def print_json_file(ld):
    fileName = "scraperData/locationData.json"
    with open(fileName, "w") as fp:
        json.dump(ld, fp, indent=4)